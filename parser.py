import requests
from bs4 import BeautifulSoup
import time
from pprint import pprint
import re
import pandas as pd
from smtplib import SMTP_SSL
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import os
import smtplib
from datetime import date, timedelta
from openpyxl import load_workbook
from pdfrw import PdfWriter
import pdfkit

URL = 'http://pub-mex.dls.gov.ua/QLA/DocList.aspx'
HEADERS = {'user-agent': 'Mzilla/5.0 (X11; Linux x86_64; rv:60.0) Gecko/20100101 Firefox/60.0', 'accept': '*/*'}
HOST = 'http://pub-mex.dls.gov.ua'

def get_html(url, params=None):
    r = requests.get(url, headers=HEADERS, params=params)
    return r

def chek(row,atrib):
    k=row.find('td',id=re.compile(atrib))
    if k:
        return k.string
def get_content(html):
    soup = BeautifulSoup(html, 'html.parser')
    headders = soup.find_all('tr', class_='rGridHeader')
    table = soup.find('table', class_='rGrid')
    dragg={'No з/п':[],
           'Дата і номер розпорядження/рішення/припису':[],
           'Дата одержання розпорядження/рішення/припису':[],
           'Назва лікарських засобів/n та перелік серій лікарських /n засобів, зазначених у розпорядженні/ рішенні/припис':[],
           'Результати перевірки щодо наявності зазначених лікарських засобів (у разі виявлення вказати кількість виявлених упаковок або зазначити: «відсутні» у разі відсутності таких':[],
           'Вжиті заходи у разі виявлення зазначених лікарських засобів*':[],
           'Підпис уповноваженої особи':[]}
    for row in table.find_all('tr',class_=re.compile('Row')):
        dragg['Назва лікарських засобів/n та перелік серій лікарських /n засобів, зазначених у розпорядженні/ рішенні/припис'].append(chek(row,'DrugName')+',Серiя №'+chek(row,'SerialNum')+','+chek(row,'Country'))
        dragg['Дата і номер розпорядження/рішення/припису'].append(chek(row,'RegDate'))
        dragg['Дата одержання розпорядження/рішення/припису'].append(date.today().strftime('%d.%m.%Y')+'\n'+chek(row,'RegNum'))
        dragg['No з/п'].append('   ')
        dragg['Результати перевірки щодо наявності зазначених лікарських засобів (у разі виявлення вказати кількість виявлених упаковок або зазначити: «відсутні» у разі відсутності таких'].append('   ') 
        if chek(row,'DocType') == 'пост. заборона ':
            postanova='знищення, повернення постачальнику'
        elif chek(row, 'DocType') == 'часткове скасув.':
            postanova='поновлення обiгу'
        elif chek(row, 'DocType') == 'тимч. заборона':
            postanova='перемiщення в карантин'
        elif chek(row,'DocType') == 'скасув. тимч. заборони':
            postanova='поновлення обiгу'
        else:
            postanova = '   '
        dragg['Вжиті заходи у разі виявлення зазначених лікарських засобів*'].append(postanova)
        dragg['Підпис уповноваженої особи'].append('   ')
    print(dragg)
    
    dd = pd.DataFrame(dragg)
    dd.to_excel('./drugs.xlsx',index=False)
    if dragg['No з/п']:
        return dragg
    else:
        return None

def parse():
    d = date.today()
    date_end=d.strftime('%d.%m.%Y')
    date_begin=(d-timedelta(days=3)).strftime('%d.%m.%Y')
    print(date_begin, date_end)
    params={'__EVENTTARGET':'ctl00$Content$fvParams$UpdateButton','ctl00$Content$fvParams$edtDocDateBegin':'{}'.format(date_begin),
            'ctl00$Content$fvParams$edtDocDateEnd':'{}'.format(date_end)}
    html = get_html(URL,params)
    if html.status_code == 200:
        print('good')
        rows = get_content(html.text)
        return rows
    else:
        print('Error')
        return None

def send_mail(fromadd,toadd,password):
    filepath ='./drugs.xlsx'
    basename = os.path.basename(filepath)
    address = toadd

    # Compose attachment
    part = MIMEBase('application', "octet-stream")
    part.set_payload(open(filepath,"rb").read() )
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', 'attachment; filename="drugs.xlsx"')

    # Compose message
    msg = MIMEMultipart()
    msg['From'] =fromadd
    msg['To'] = toadd
    msg.attach(part)

    # Send mail
    server = smtplib.SMTP_SSL('smtp.gmail.com',465)
    server.connect('smtp.gmail.com',465)
    server.login(fromadd,password)
    server.sendmail(msg['From'], msg['To'], msg.as_string())
    server.quit()

def convert_to_pdf(exelfile,pdfile):
    df = pd.read_excel(exelfile)#input
    df.to_html("file.html")#to html
    pdfkit.from_file("file.html", pdfile)#to pdf

def convert_to_pdf2(exelfile,pdfile):
    wb = pd.read_excel(exelfile)
    wb=PdfWriter()
    wb.write(pdfile)

def convert_to_pdf3(exelfile,pdfile):
    workbook = load_workbook(exelfile, data_only=True)
    worksheet = workbook.active
    pw = PdfWriter(pdfile)
    #pw.setFont('Courier', 12)
    #pw.setHeader('Журнал бла бла бла')
    #pw.setFooter('Generated using openpyxl and xtopdf')

    ws_range = worksheet.iter_rows('A1:H13')
    for row in ws_range:
        s = ''
        for cell in row:
            if cell.value is None:
                s += ' ' * 11
            else:
                s += str(cell.value).rjust(10) + ' '
                pw.writeLine(s)
                pw.savePage()
                pw.close()
def convert_to_pdf4(exelfile,pdfile):
    o = win32com.client.Dispatch("Excel.Application")
    o.Visible = False
    o.DisplayAlerts = False
    wb = o.Workbooks.Open(exelfile,pdfile)
    wb.WorkSheets("sheet1").Select()
    wb.ActiveSheet.ExportAsFixedFormat(0,pdfile)
    o.Quit()

if __name__=='__main__':
    if parse():
        send_mail()




